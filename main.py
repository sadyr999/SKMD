import openpyxl
import shutil
import datetime
import glob
import os
import pandas
import random
import json
import pandas as pd
from pandasql import sqldf
import math
from collections import defaultdict, OrderedDict
def backupBase():
    print("Backing up BASE..")
    randomNum = random.randrange(1, 99)
    shutil.copy(pathBase, f"arhvBKP/BKPbase{currStamp}-{str(randomNum)}.xlsx")
def backupLog():
    print("Backing up LOG..")
    randomNum = random.randrange(1, 99)
    shutil.copy("LogReceptions.xlsx", f"arhvBKP/BKPLog{currStamp}-{str(randomNum)}.xlsx")
def dailyImport():
    backupBase()
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    lastRawBase = sheetBase.max_row + 1

    for file in glob.glob("Страх*.xlsx"):
        if file in open(pathLog).read():
            print(file + " - DUPLICATE")
            os.remove(file)
            pass
        else:
            with open(pathLog, "a+") as f:
                pathIns = file
                insReport = file
                bookIns = openpyxl.load_workbook(pathIns)
                sheetIns = bookIns.active
                bookNew = openpyxl.load_workbook(pathRepINS)
                sheetNew = bookNew.active

                maxRow = sheetIns.max_row
                maxColumn = sheetIns.max_column

                counterA = 4
                # while counterA < maxRow:
                while sheetIns.cell(row=counterA, column=2).value != None:

                    dateS = str(sheetIns.cell(row=counterA, column=6).value)
                    startDate = dateS[6:10] + "-" + dateS[3:5] + "-" + dateS[0:2]

                    sheetNew.cell(row=counterA, column=1).value = sheetIns.cell(row=counterA, column=12).value
                    sheetNew.cell(row=counterA, column=2).value = sheetIns.cell(row=counterA, column=13).value
                    sheetNew.cell(row=counterA, column=3).value = sheetIns.cell(row=counterA, column=4).value
                    sheetNew.cell(row=counterA, column=4).value = sheetIns.cell(row=counterA,
                                                                                column=1).value + " " + str(
                        sheetIns.cell(row=counterA, column=2).value)
                    sheetNew.cell(row=counterA, column=5).value = str(sheetIns.cell(row=counterA, column=6).value)[0:10]
                    sheetNew.cell(row=counterA, column=6).value = sheetIns.cell(row=counterA, column=7).value.strftime(
                        "%d.%m.%Y")
                    sheetNew.cell(row=counterA, column=7).value = sheetIns.cell(row=counterA, column=10).value

                    sheetNew.cell(row=counterA, column=8).value = math.ceil(
                        sheetIns.cell(row=counterA, column=14).value / 30)
                    sheetNew.cell(row=counterA, column=9).value = round(sheetIns.cell(row=counterA, column=15).value,
                                                                        1) - 0.1
                    sheetNew.cell(row=counterA, column=10).value = sheetIns.cell(row=counterA,
                                                                                 column=10).value * sheetNew.cell(
                        row=counterA, column=9).value / 100
                    sheetNew.cell(row=counterA, column=11).value = sheetNew.cell(row=counterA, column=10).value * 0.03

                    sheetNew.cell(row=counterA, column=12).value = sheetIns.cell(row=counterA, column=3).value

                    phoneCell = ""
                    phoneOrig = str(sheetIns.cell(row=counterA, column=18).value)[0:9]
                    phoneRaw = ""
                    for i in phoneOrig:
                        if i.isnumeric():
                            phoneRaw += i
                    phoneRaw = "000000000" + phoneRaw
                    phoneCell = "996" + phoneRaw[-9:]

                    sheetNew.cell(row=counterA, column=13).value = phoneCell

                    for i in range(1, maxColumn + 1):
                        sheetBase.cell(row=lastRawBase, column=i).value = sheetNew.cell(row=counterA, column=i).value

                    lastRawBase += 1
                    counterA += 1
                f.write(f"{file}\n")
                f.close()
                bookNew.save(f"aaImport/2XX 00-00 {pathIns}")
                destination_path = "arhvIns/recep" + file
                shutil.move(file, destination_path)
                print(f"MD Import File - {pathIns} - Processed")
    bookBase.save(pathBase)
    checkUniqueClient()
def updateCallsBase(file: str) -> str:
    print("Backing up CALLS..")
    shutil.copy(pathEmer, f"arhvBKP/BKPemer{currStamp}.xlsx")
    shutil.copy(pathPlan, f"arhvBKP/BKPplan{currStamp}.xlsx")
    shutil.copy(pathAll, f"arhvBKP/BKPall{currStamp}.xlsx")

    bookEmer = openpyxl.load_workbook(pathEmer)
    sheetEmer = bookEmer.active
    bookPlan = openpyxl.load_workbook(pathPlan)
    sheetPlan = bookPlan.active
    bookImport = openpyxl.load_workbook(file)
    sheetImport = bookImport.active
    bookAll = openpyxl.load_workbook(pathAll)
    sheetAll = bookAll.active

    display = 0
    lastRowEmer = sheetEmer.max_row + 1
    lastRowPlan = sheetPlan.max_row + 1
    lastRowAll = sheetAll.max_row + 1
    lastRowImport = sheetImport.max_row
    print("Processing Receptions File ", end="")
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    bookBaseNew = openpyxl.load_workbook(pathBaseNew)
    sheetBaseNew = bookBaseNew.active
    counterRow = lastRowImport
    print("+base ", end="")
    maxRowOR = sheetBase.max_row
    maxRowNewOR = sheetBaseNew.max_row

    while counterRow > 1:
        maxRow = maxRowOR
        maxRowNew = maxRowNewOR
        policyNum = "НЕТ ПОЛИСА"
        officeName = "НЕТ ПОЛИСА"
        callcenter = ""
        phone = ""
        status = ""
        dateReg = ""
        agentID = ""
        medcard = str(sheetImport.cell(row=counterRow, column=14).value) + " - " + str(
            sheetImport.cell(row=counterRow, column=15).value) + " - " + str(
            sheetImport.cell(row=counterRow, column=16).value) + " - " + str(
            sheetImport.cell(row=counterRow, column=17).value)
        nameIns = sheetImport.cell(row=counterRow, column=5).value
        notNew = True
        noRecord = True

        while maxRowNew > 1:
            if sheetBaseNew.cell(row=maxRowNew, column=2).value.strip() == nameIns.strip():
                agentID = sheetBaseNew.cell(row=maxRowNew, column=1).value
                callcenter = sheetBaseNew.cell(row=maxRowNew, column=6).value
                phone = sheetBaseNew.cell(row=maxRowNew, column=3).value
                notNew = False
                noRecord = False
                maxRowNew = 0
            maxRowNew -= 1

        while notNew and maxRow > 1:
            if sheetBase.cell(row=maxRow, column=3).value == nameIns:
                policyNum = sheetBase.cell(row=maxRow, column=4).value
                officeName = sheetBase.cell(row=maxRow, column=2).value
                callcenter = sheetBase.cell(row=maxRow, column=25).value
                phone = sheetBase.cell(row=maxRow, column=13).value
                maxRow = 0
                noRecord = False
            maxRow -= 1

        # if len(sheetImport.cell(row=counterRow, column=2).value) < 10:
        if "деж" not in (sheetImport.cell(row=counterRow, column=3).value):
            sheetPlan.cell(row=lastRowPlan, column=1).value = str(sheetImport.cell(row=counterRow, column=1).value)[:10]
            for i in range(2, 6):
                sheetPlan.cell(row=lastRowPlan, column=i).value = sheetImport.cell(row=counterRow, column=i).value
            doctor = str(sheetImport.cell(row=counterRow, column=3).value)
            if doctor in open(pathSov).read():
                sheetPlan.cell(row=lastRowPlan, column=6).value = "Медицинский Советник"
            elif doctor in open(pathOrg).read():
                sheetPlan.cell(row=lastRowPlan, column=6).value = "Организация лечения"
            else:
                sheetPlan.cell(row=lastRowPlan, column=6).value = "Специалист"
            sheetPlan.cell(row=lastRowPlan, column=7).value = policyNum
            sheetPlan.cell(row=lastRowPlan, column=8).value = officeName
            for i in range(1, 10):
                sheetAll.cell(row=lastRowAll, column=i).value = sheetPlan.cell(row=lastRowPlan, column=i).value
            lastRowPlan += 1
        else:
            sheetEmer.cell(row=lastRowEmer, column=1).value = str(sheetImport.cell(row=counterRow, column=1).value)[:10]
            for i in range(2, 6):
                sheetEmer.cell(row=lastRowEmer, column=i).value = sheetImport.cell(row=counterRow, column=i).value
            topic = str(sheetImport.cell(row=counterRow, column=6).value)
            if "не" in topic:
                sheetEmer.cell(row=lastRowEmer, column=6).value = "Дежурный врач"
            else:
                sheetEmer.cell(row=lastRowEmer, column=6).value = "Дежурный врач с назначением лечения"
            sheetEmer.cell(row=lastRowEmer, column=7).value = policyNum
            sheetEmer.cell(row=lastRowEmer, column=8).value = officeName
            for i in range(1, 10):
                sheetAll.cell(row=lastRowAll, column=i).value = sheetEmer.cell(row=lastRowEmer, column=i).value
            lastRowEmer += 1

        sheetAll.cell(row=lastRowAll, column=7).value = policyNum
        sheetAll.cell(row=lastRowAll, column=8).value = officeName
        sheetAll.cell(row=lastRowAll, column=16).value = agentID
        sheetAll.cell(row=lastRowAll, column=17).value = callcenter
        sheetAll.cell(row=lastRowAll, column=15).value = phone
        sheetAll.cell(row=lastRowAll, column=25).value = medcard
        sheetAll.cell(row=lastRowAll, column=24).value = sheetImport.cell(row=counterRow, column=13).value

        sheetAll.cell(row=lastRowAll, column=18).value = sheetImport.cell(row=counterRow, column=11).value
        sheetAll.cell(row=lastRowAll, column=19).value = sheetImport.cell(row=counterRow, column=9).value
        sheetAll.cell(row=lastRowAll, column=20).value = sheetImport.cell(row=counterRow, column=8).value

        counterPolicy = lastRowAll - 1
        policyNew = True
        clientNew = True
        quantityEmer = 0
        quantityPlan = 0

        while counterPolicy > 2:
            if sheetAll.cell(row=lastRowAll, column=5).value == sheetAll.cell(row=counterPolicy, column=5).value:
                # if sheetAll.cell(row=counterPolicy, column=7).value == sheetAll.cell(row=lastRowAll, column=7).value \
                #        and sheetAll.cell(row=counterPolicy, column=7).value != "НЕТ ПОЛИСА":
                if sheetAll.cell(row=counterPolicy, column=7).value != "НЕТ ПОЛИСА":
                    sheetAll.cell(row=lastRowAll, column=9).value = "повторное"
                    sheetAll.cell(row=lastRowAll, column=10).value = sheetAll.cell(row=counterPolicy, column=10).value
                    policyNew = False
                clientNew = False
                if "Дежурный" in sheetAll.cell(row=lastRowAll, column=6).value:
                    sheetAll.cell(row=lastRowAll, column=11).value = "дежурный"
                    quantityPlan = int(sheetAll.cell(row=counterPolicy, column=12).value)
                    quantityEmer = int(sheetAll.cell(row=counterPolicy, column=13).value) + 1
                else:
                    sheetAll.cell(row=lastRowAll, column=11).value = "плановый"
                    quantityPlan = int(sheetAll.cell(row=counterPolicy, column=12).value) + 1
                    quantityEmer = int(sheetAll.cell(row=counterPolicy, column=13).value)
                counterPolicy = 0
                break
            counterPolicy -= 1

        if policyNew:
            if clientNew:
                if "Дежурный" in sheetAll.cell(row=lastRowAll, column=6).value:
                    sheetAll.cell(row=lastRowAll, column=11).value = "дежурный"
                    quantityEmer = 1
                else:
                    sheetAll.cell(row=lastRowAll, column=11).value = "плановый"
                    quantityPlan = 1
            if sheetAll.cell(row=lastRowAll, column=7).value != "НЕТ ПОЛИСА":
                sheetAll.cell(row=lastRowAll, column=9).value = "открытие файла"
                sheetAll.cell(row=lastRowAll, column=10).value = policyNum + " от " + sheetAll.cell(row=lastRowAll,
                                                                                                    column=1).value
        sheetAll.cell(row=lastRowAll, column=12).value = quantityPlan
        sheetAll.cell(row=lastRowAll, column=13).value = quantityEmer
        sheetAll.cell(row=lastRowAll, column=14).value = quantityEmer + quantityPlan

        lastRowAll += 1
        counterRow -= 1

        if (lastRowImport - counterRow) / lastRowImport * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 5

    print("")
    print("Saving..")
    bookEmer.save(pathEmer)
    bookPlan.save(pathPlan)
    bookAll.save(pathAll)

    rangeReturn = str(sheetImport.cell(row=2, column=1).value)[:10] + "-" + str(
        sheetImport.cell(row=sheetImport.max_row, column=1).value)[:10]
    return rangeReturn
def TelegramScheldule(path_json):
    f = open(path_json, encoding="utf8'")
    data = json.load(f)

    with open('resourses/file.json', 'w', encoding='utf8') as f:
        # for message in data['messages']:
        #     json.dump(message, f, indent=2)

        new_data = {'data': data['messages']}
        json.dump(new_data, f, indent=2, ensure_ascii=False)

        values = new_data['data']
        plan_list = {"data": []}
        new_list = {"data": []}
        doc_con_count = 0
        pac_con_count = 0
        for value in values:
            try:
                if type(value["text"][0]) == dict:
                    index_of_text = 0
                elif type(value["text"][0]) == str:
                    index_of_text = 1

                time = value["text"][index_of_text + 2]["text"]
                specialist = value["text"][index_of_text + 8]["text"]
                patient = value["text"][index_of_text + 4]["text"]
                text = value["text"][index_of_text]["text"]
                phone_number = value["text"][index_of_text + 6]["text"]
                date_str = value["date"].replace('T', ' ')
                format = "%Y-%m-%d %H:%M:%S"
                date = datetime.datetime.strptime(date_str, format).date()
                date_time = datetime.datetime.strptime(date_str, format).time()

                if "Запланированный прием" in text:
                    if len(value["text"]) <= 12:
                        new_val_i = {
                            "дата": str(date),
                            "время": time,
                            "специалист": specialist,
                            "пациент": patient,
                            "номер пациента": str(phone_number),
                            "врач подключился": "",
                            "пациент подключился": "",
                            "врач отключился": "",
                            "пациент отключился": ""
                        }
                        plan_list["data"].append(new_val_i)
                    elif len(value["text"]) > 12:
                        new_val_i = {
                            "дата": str(date),
                            "время": time,
                            "специалист": specialist,
                            "пациент": patient,
                            "номер пациента": str(phone_number),
                            "врач подключился": "",
                            "пациент подключился": "",
                            "врач отключился": "",
                            "пациент отключился": ""
                        }
                        plan_list["data"].append(new_val_i)

                        if index_of_text == 0:
                            range_i = (len(value["text"]) - 2) // 10
                        elif index_of_text == 1:
                            range_i = (len(value["text"]) - 3) // 10

                        for i in range(1, range_i):
                            index_of_value = i * 10 + index_of_text
                            patient = value["text"][index_of_value + 4]["text"]
                            specialist = value["text"][index_of_value + 8]["text"]
                            phone_number = value["text"][index_of_value + 6]["text"]
                            new_val_i = {
                                "дата": str(date),
                                "время": time,
                                "специалист": specialist,
                                "пациент": patient,
                                "номер пациента": str(phone_number),
                                "врач подключился": "",
                                "пациент подключился": "",
                                "врач отключился": "",
                                "пациент отключился": ""
                            }
                            plan_list["data"].append(new_val_i)

                if text == "Специалист подключился к приему":
                    for dict_i in plan_list["data"]:
                        if dict_i["специалист"] == specialist and dict_i["пациент"] == patient and dict_i[
                            "время"] == time and dict_i["дата"] == str(date):
                            dict_i["врач подключился"] = str(date_time)
                            # new_list['data'].append(dict_i)
                    doc_con_count += 1
                if text == "Пациент подключился к приему":
                    for dict_i in plan_list["data"]:
                        if dict_i["специалист"] == specialist and dict_i["пациент"] == patient and dict_i[
                            "время"] == time and dict_i["дата"] == str(date):
                            dict_i["пациент подключился"] = str(date_time)
                    pac_con_count += 1
                if text == "Специалист отключился от приема":
                    for dict_i in plan_list["data"]:
                        if dict_i["специалист"] == specialist and dict_i["пациент"] == patient and dict_i[
                            "время"] == time and dict_i["дата"] == str(date):
                            dict_i["врач отключился"] = str(date_time)
                if text == "Пациент отключился от приема":
                    for dict_i in plan_list["data"]:
                        if dict_i["специалист"] == specialist and dict_i["пациент"] == patient and dict_i[
                            "время"] == time and dict_i["дата"] == str(date):
                            dict_i["пациент отключился"] = str(date_time)
                if text == "Специалист перезванивает пациенту":
                    for dict_i in plan_list["data"]:
                        if dict_i["специалист"] == specialist and dict_i["пациент"] == patient and dict_i[
                            "время"] == time and dict_i["дата"] == str(date):
                            dict_i["врач перезвонил"] = str(date_time)
            except Exception as ex:
                pass

        with open('resourses/data.json', 'w', encoding='utf8') as f:
            print(str(len(plan_list["data"])), " Запланированных приемов")
            print(str(doc_con_count), " Докторов подключились")
            print(str(pac_con_count), " Пациентов подключились")
            json.dump(plan_list, f, indent=2, ensure_ascii=False)
            # print("New json file is created from data.json file")
    name_xlsx = 'плановые_приемы.xlsx'
    try:
        os.remove(name_xlsx)
    except Exception:
        pass
    pandas.read_json("resourses/data.json", orient='split').to_excel(name_xlsx, index=False)
    print("Создана exel таблица с плановыми приемами")
def TelegramCall(path_json):
    f = open(path_json, encoding="utf8'")
    data = json.load(f)

    with open('resourses/file_call.json', 'w', encoding='utf8') as f:
        # for message in data['messages']:
        #     json.dump(message, f, indent=2)

        new_data = {'data': data['messages']}
        json.dump(new_data, f, indent=2, ensure_ascii=False)

        values = new_data['data']
        plan_list = {"data": []}
        # new_list = {"data": []}
        # doc_con_count = 0
        # pac_con_count = 0
        for value in values:
            try:
                text = value["text"][1]
                date_str = value["date"].replace('T', ' ')
                format = "%Y-%m-%d %H:%M:%S"
                date = datetime.datetime.strptime(date_str, format).date()
                time = datetime.datetime.strptime(date_str, format).time()
                date_time = datetime.datetime.strptime(date_str, format).time()
                if type(text) == str:
                    if "❓\nЗвонок от " in text:
                        patient = value["text"][2]["text"]
                        phone_number = value["text"][4]["text"]
                        new_val_i = {
                            "дата": str(date),
                            "время": str(time),
                            "пациент": patient,
                            "номер пациента": str(phone_number),
                            "принял специалист": "",
                        }
                        plan_list["data"].append(new_val_i)
                    elif "Пациент -" in text:
                        text = value["text"][0]["text"]
                        if text == "Пациент вышел из очереди на дежурный прием":
                            patient = value["text"][2]["text"]
                            for dict_i in plan_list["data"]:
                                if dict_i["пациент"] == patient and dict_i["дата"] == str(date) and dict_i[
                                    "принял специалист"] == "":
                                    dict_i["принял специалист"] = "вышел"
                    else:
                        text = value["text"][3]
                        if "начал созвон с пользователем" in text:
                            specialist = value["text"][2]["text"]
                            patient = value["text"][4]["text"]
                            for dict_i in plan_list["data"]:
                                if dict_i["пациент"] == patient and dict_i["дата"] == str(date) and dict_i[
                                    "принял специалист"] == "":
                                    dict_i["принял специалист"] = specialist
            except Exception as ex:
                pass

        with open('resourses/data_call.json', 'w', encoding='utf8') as f:
            print(str(len(plan_list["data"])), " Дежурных вызовов")
            json.dump(plan_list, f, indent=2, ensure_ascii=False)
            # print("New json file is created from data.json file")
    name_xlsx = 'дежурные_вызовы.xlsx'
    try:
        os.remove(name_xlsx)
    except Exception:
        pass
    pandas.read_json("resourses/data_call.json", orient='split').to_excel(name_xlsx, index=False)
    print("Создана exel таблица с дежурными вызовами")
def TelegramEntries(path_json):
    f = open(path_json, encoding="utf8'")
    data = json.load(f)

    with open('resourses/file_record.json', 'w', encoding='utf8') as f:
        new_data = {'data': data['messages']}
        json.dump(new_data, f, indent=2, ensure_ascii=False)

        values = new_data['data']
        plan_list = {"data": []}
        index_of_text = 0
        for value in values:
            try:
                if type(value["text"][0]) == dict:
                    index_of_text = 0
                elif type(value["text"][0]) == str:
                    index_of_text = 1
                try:
                    text = value["text"][index_of_text]["text"]
                    if text.find("Новая запись на прием") != -1:

                        date_str = value["text"][index_of_text + 2]["text"]
                        format = "%Y-%m-%d %H:%M"
                        date = datetime.datetime.strptime(date_str, format).date()
                        time = datetime.datetime.strptime(date_str, format).time()
                        specialist = value["text"][index_of_text + 8]["text"]
                        patient = value["text"][index_of_text + 4]["text"]
                        phone_number = value["text"][index_of_text + 6]["text"]

                        index_open_bracket = int(text.find("("))
                        index_close_bracket = int(text.find(")"))
                        if index_open_bracket != -1 and index_close_bracket != -1:
                            author = text[index_open_bracket + 1:index_close_bracket]
                        else:
                            author = ""
                        new_val_i = {
                            "дата": str(date),
                            "время": str(time),
                            "специалист": specialist,
                            "пациент": patient,
                            "запись": author,
                            "номер пациента": str(phone_number),
                        }
                        plan_list["data"].append(new_val_i)
                except Exception as ex:
                    pass
            except Exception as ex:
                pass

        with open('resourses/data_record.json', 'w', encoding='utf8') as f:
            print(str(len(plan_list["data"])), " Записей")
            json.dump(plan_list, f, indent=2, ensure_ascii=False)
            # print("New json file is created from data.json file")
    name_xlsx = 'записи.xlsx'
    try:
        os.remove(name_xlsx)
    except Exception:
        pass
    pandas.read_json("resourses/data_record.json", orient='split').to_excel(name_xlsx, index=False)
    print("Создана exel таблица с записями")
def UpdateAgentList():
    print("Backing up BASE..")
    randomNum = random.randrange(1, 99)
    shutil.copy(pathBaseNew, f"arhvBKP/BKPbaseNew{currStamp}-{str(randomNum)}.xlsx")
    display = 0
    print("Processing EMER-RECORD ", end="")
    bookBaseNew = openpyxl.load_workbook(pathBaseNew)
    sheetBaseNew = bookBaseNew.active
    bookEntries = openpyxl.load_workbook('записи.xlsx')
    sheetEntries = bookEntries.active
    bookCalls = openpyxl.load_workbook('дежурные_вызовы.xlsx')
    sheetCalls = bookCalls.active

    maxRowClients = sheetBaseNew.max_row
    maxRowEntries = sheetEntries.max_row
    maxRowCalls = sheetCalls.max_row
    counterClients = 2

    while counterClients <= maxRowClients:
        counterCalls = 2
        counterEntries = 2
        nameClient = sheetBaseNew.cell(row=counterClients, column=2).value.strip()
        sheetBaseNew.cell(row=counterClients, column=9).value = "neakt"
        while counterCalls < maxRowCalls:
            if nameClient == sheetCalls.cell(row=counterCalls, column=3).value.strip():
                sheetBaseNew.cell(row=counterClients, column=7).value = "yes"
                sheetBaseNew.cell(row=counterClients, column=9).value = "aktiv"
                break
            counterCalls += 1
        while counterEntries < maxRowEntries:
            if nameClient == sheetEntries.cell(row=counterEntries, column=4).value.strip():
                sheetBaseNew.cell(row=counterClients, column=8).value = "yes"
                sheetBaseNew.cell(row=counterClients, column=9).value = "aktiv"
                break
            counterEntries += 1
        if counterClients / maxRowClients * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 5
        counterClients += 1
    print("")
    bookBaseNew.save(pathBaseNew)
def TelegramCallCC():
    name_xlsx = 'дежурные_вызовы.xlsx'
    name_xlsx2 = 'записи.xlsx'
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    bookCalls = openpyxl.load_workbook(name_xlsx)
    sheetCalls = bookCalls.active
    bookRecords = openpyxl.load_workbook(name_xlsx2)
    sheetRecords = bookRecords.active
    counterRow = 2
    maxRowCalls = sheetCalls.max_row
    maxRowRecords1 = sheetRecords.max_row
    maxRow1 = sheetBase.max_row
    display = 0
    print("Processing EMER Calls ", end="")
    while counterRow <= maxRowCalls:
        nameIns = str(sheetCalls.cell(row=counterRow, column=4).value)
        maxRow = maxRow1
        maxRowRecords = maxRowRecords1
        haveRecord = ""
        callcenter = ""
        policy = ""
        beginP = ""
        weekC = ""
        while maxRow > 1:
            if nameIns == sheetBase.cell(row=maxRow, column=13).value and sheetBase.cell(row=maxRow,
                                                                                         column=25).value != None:
                callcenter = sheetBase.cell(row=maxRow, column=25).value
                policy = sheetBase.cell(row=maxRow, column=4).value
                beginP = sheetBase.cell(row=maxRow, column=5).value
                weekC = sheetBase.cell(row=maxRow, column=24).value

                sheetCalls.cell(row=counterRow, column=6).value = callcenter
                sheetCalls.cell(row=counterRow, column=7).value = policy
                sheetCalls.cell(row=counterRow, column=8).value = beginP
                sheetCalls.cell(row=counterRow, column=9).value = weekC

                break
                print("wwwww")
                maxRow = 0
            maxRow -= 1

        while maxRowRecords > 1:
            if str(sheetRecords.cell(row=maxRowRecords, column=6).value) == nameIns:
                haveRecord = "есть"
                sheetCalls.cell(row=counterRow, column=11).value = haveRecord
                break
            maxRowRecords -= 1

        my_date = datetime.datetime.strptime(sheetCalls.cell(row=counterRow, column=1).value, "%Y-%m-%d")
        year, week_num, day_of_week = my_date.isocalendar()
        currWeek = week_num
        sheetCalls.cell(row=counterRow, column=10).value = currWeek

        counterRow += 1
        if counterRow / maxRowCalls * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 5
    print("")
    print("CC Processed")
    bookCalls.save(name_xlsx)
def TelegramPlanCC():
    name_xlsx = 'плановые_приемы.xlsx'
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    bookCalls = openpyxl.load_workbook(name_xlsx)
    sheetCalls = bookCalls.active
    counterRow = 2
    maxRowCalls = sheetCalls.max_row
    maxRow1 = sheetBase.max_row
    display = 0
    print("Processing Receptions File ", end="")
    while counterRow <= maxRowCalls:
        nameIns = sheetCalls.cell(row=counterRow, column=4).value
        maxRow = maxRow1
        callcenter = ""
        policy = ""
        beginP = ""
        weekC = ""
        while maxRow > 1:
            if sheetBase.cell(row=maxRow, column=3).value == nameIns:
                callcenter = sheetBase.cell(row=maxRow, column=25).value
                policy = sheetBase.cell(row=maxRow, column=4).value
                beginP = sheetBase.cell(row=maxRow, column=5).value
                weekC = sheetBase.cell(row=maxRow, column=24).value
                maxRow = 0
            maxRow -= 1
        sheetCalls.cell(row=counterRow, column=11).value = callcenter
        sheetCalls.cell(row=counterRow, column=12).value = policy
        sheetCalls.cell(row=counterRow, column=13).value = beginP
        sheetCalls.cell(row=counterRow, column=14).value = weekC

        my_date = datetime.datetime.strptime(sheetCalls.cell(row=counterRow, column=1).value, "%Y-%m-%d")
        year, week_num, day_of_week = my_date.isocalendar()
        currWeek = week_num
        sheetCalls.cell(row=counterRow, column=15).value = currWeek

        counterRow += 1
        if counterRow / maxRowCalls * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 5
    print("")
    print("CC Processed")
    bookCalls.save(name_xlsx)
def smsReportImport():
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    bookPRCD = openpyxl.load_workbook(pathPRCD)
    sheetPRCD = bookPRCD.active

    for file in glob.glob("Отчет за период*.xlsx"):
        bookSMS = openpyxl.load_workbook(file)
        sheetSMS = bookSMS.active

        sheetSMS.delete_cols(10)
        sheetSMS.delete_cols(9)
        sheetSMS.delete_cols(8)
        sheetSMS.delete_cols(7)
        sheetSMS.delete_cols(6)
        sheetSMS.delete_cols(5)
        sheetSMS.delete_cols(4)
        sheetSMS.delete_cols(2)

        maxRowSMS = sheetSMS.max_row
        maxRowBase = sheetBase.max_row
        maxRowPRCD = sheetPRCD.max_row
        print("Processing SMS Report ", end="")
        display = 0
        counterA = maxRowSMS
        while counterA > 1:
            phoneCell = sheetSMS.cell(row=counterA, column=2).value
            counterB = counterA - 1
            while counterB > 0:
                if phoneCell == sheetSMS.cell(row=counterB, column=2).value:
                    sheetSMS.delete_rows(counterB)
                    counterA -= 1
                counterB -= 1
            counterC = 1
            while counterC <= maxRowBase:
                if phoneCell == sheetBase.cell(row=counterC, column=13).value:
                    sheetSMS.cell(row=counterA, column=3).value = sheetBase.cell(row=counterC, column=3).value
                    sheetSMS.cell(row=counterA, column=4).value = sheetBase.cell(row=counterC, column=4).value
                    sheetSMS.cell(row=counterA, column=5).value = sheetBase.cell(row=counterC, column=5).value
                    sheetSMS.cell(row=counterA, column=6).value = sheetBase.cell(row=counterC, column=2).value
                    sheetSMS.cell(row=counterA, column=7).value = sheetBase.cell(row=counterC, column=1).value
                    sheetSMS.cell(row=counterA, column=8).value = sheetBase.cell(row=counterC, column=12).value
                    break
                counterC += 1
                if counterC > maxRowBase:
                    sheetSMS.delete_rows(counterA)
            counterA -= 1
            if (maxRowSMS - counterA) / maxRowSMS * 100 > display:
                print(">" + str(display) + "%", end="")
                display += 10
        maxRowSMS = sheetSMS.max_row
        counterA = maxRowSMS
        while counterA > 1:
            counterB = maxRowPRCD
            while counterB > 1:
                if sheetSMS.cell(row=counterA, column=2).value == sheetPRCD.cell(row=counterB, column=2).value:
                    counterB = 0
                    break
                if counterB == 2:
                    for i in range(1, 9):
                        sheetPRCD.cell(row=maxRowPRCD + 1, column=i).value = sheetSMS.cell(row=counterA, column=i).value
                    maxRowPRCD += 1
                counterB -= 1
            counterA -= 1
        bookPRCD.save(pathPRCD)
        destination_path = "arhvSMS/" + file
        shutil.move(file, destination_path)
        print("")
        print(f"SMS Report - {file} - Processed")

    print("Forming BF Report.. ")
    bookCalls = openpyxl.load_workbook(pathAll)
    sheetCalls = bookCalls.active

    while True:
        print("00 for EXIT")
        d1 = str(input("Day 1 of the week / mm-dd : "))
        if d1 == "00":
            break
        d2 = str(input("Day 2 of the week / mm-dd : "))
        d3 = str(input("Day 3 of the week / mm-dd : "))
        d4 = str(input("Day 4 of the week / mm-dd : "))
        d5 = str(input("Day 5 of the week / mm-dd : "))
        d6 = str(input("Day 6 of the week / mm-dd : "))
        d7 = str(input("Day 7 of the week / mm-dd : "))
        if d2 == "00" or d3 == "00" or d4 == "00" or d5 == "00":
            break
        s1 = d1[3] + d1[4] + "." + d1[0] + d1[1]
        s2 = d2[3] + d2[4] + "." + d2[0] + d2[1]
        s3 = d3[3] + d3[4] + "." + d3[0] + d3[1]
        s4 = d4[3] + d4[4] + "." + d4[0] + d4[1]
        s5 = d5[3] + d5[4] + "." + d5[0] + d5[1]
        s6 = d6[3] + d6[4] + "." + d6[0] + d6[1]
        s7 = d7[3] + d7[4] + "." + d7[0] + d7[1]

        bookReport = openpyxl.load_workbook("TemplateReport.xlsx")
        sheetReport = bookReport.active

        maxRowSMS = sheetPRCD.max_row
        maxRowBase = sheetBase.max_row
        maxRowReport = 51  # sheetReport.max_row - 2
        maxRowCalls = sheetCalls.max_row

        # move all data
        counterFcol = 27
        counterFrow = 2
        for i in range(1, 12):
            if True:
                while counterFrow <= 54:
                    for j in range(0, 2):
                        if counterFrow != 51:
                            sheetReport.cell(row=counterFrow, column=counterFcol + 2 + j).value = \
                                sheetReport.cell(row=counterFrow, column=counterFcol + j).value
                    counterFrow += 1
            counterFcol = counterFcol - 2
            counterFrow = 2

        dateEnd = sheetPRCD.cell(row=2, column=1).value
        counterD = 4
        sheetReport.cell(row=1, column=6).value = "по " + str(dateEnd)[:9]
        sheetReport.cell(row=2, column=7).value = "неделя " + d1 + " по " + d7
        femaleWeekReg = 0
        maleWeekReg = 0
        femaleWeekIns = 0
        maleWeekIns = 0
        femaleTotalReg = 0
        maleTotalReg = 0
        femaleTotalIns = 0
        maleTotalIns = 0

        print("Processing Report... ", end="")
        display = 0
        while counterD < maxRowReport:
            if counterD / maxRowReport * 100 > display:
                print(">" + str(display) + "%", end="")
                display += 3
            quantityTotal = 0
            quantityRegis = 0
            quantityTotalLastW = 0
            quantityRegisLastW = 0
            quantityCalls = 0
            quantityCallsLastW = 0
            counterE = 2
            counterF = 2
            officeName = sheetReport.cell(row=counterD, column=2).value

            # print("Counting new members..")
            while counterE <= maxRowBase:
                if officeName == sheetBase.cell(row=counterE, column=2).value:
                    if sheetBase.cell(row=counterE, column=20).value == "uniq":
                        quantityTotal += 1
                        if str(sheetBase.cell(row=counterE, column=12).value)[0] == "1":
                            femaleTotalIns += 1
                        else:
                            maleTotalIns += 1
                        dayAndMonth = str(sheetBase.cell(row=counterE, column=5).value)[-5:]
                        if d1 in dayAndMonth or d2 in dayAndMonth or d3 in dayAndMonth or \
                                d4 in dayAndMonth or d5 in dayAndMonth or d6 in dayAndMonth or d7 in dayAndMonth:
                            quantityTotalLastW += 1
                            if str(sheetBase.cell(row=counterE, column=12).value)[0] == "1":
                                femaleWeekIns += 1
                            else:
                                maleWeekIns += 1
                counterE += 1
            # print("Counting new registrations..")
            while counterF <= maxRowSMS:
                if officeName == sheetPRCD.cell(row=counterF, column=6).value:
                    quantityRegis += 1
                    if str(sheetPRCD.cell(row=counterF, column=8).value)[0] == "1":
                        femaleTotalReg += 1
                    else:
                        maleTotalReg += 1
                    dayAndMonth = str(sheetPRCD.cell(row=counterF, column=1).value)[:5]
                    if s1 in dayAndMonth or s2 in dayAndMonth or s3 in dayAndMonth or s4 in dayAndMonth or s5 in dayAndMonth or s6 in dayAndMonth or s7 in dayAndMonth:
                        quantityRegisLastW += 1
                        if str(sheetPRCD.cell(row=counterF, column=8).value)[0] == "1":
                            femaleWeekReg += 1
                        else:
                            maleWeekReg += 1
                counterF += 1
            # print("Counting receptions..")
            counterF = 2
            while counterF <= maxRowCalls:
                if officeName == sheetCalls.cell(row=counterF, column=8).value:
                    quantityCalls += 1
                    dayAndMonth = str(sheetCalls.cell(row=counterF, column=1).value)[-5:]
                    if d1 in dayAndMonth or d2 in dayAndMonth or d3 in dayAndMonth or \
                            d4 in dayAndMonth or d5 in dayAndMonth or d6 in dayAndMonth or d7 in dayAndMonth:
                        quantityCallsLastW += 1
                counterF += 1
            sheetReport.cell(row=counterD, column=3).value = quantityTotal
            sheetReport.cell(row=counterD, column=4).value = quantityRegis
            # sheetReport.cell(row=counterD, column=7).value = quantityTotalLastW
            sheetReport.cell(row=counterD, column=7).value = quantityRegisLastW
            sheetReport.cell(row=counterD, column=6).value = quantityCalls
            sheetReport.cell(row=counterD, column=8).value = quantityCallsLastW
            counterD += 1

        sheetReport.cell(row=maxRowReport + 1, column=3).value = maleTotalIns
        sheetReport.cell(row=maxRowReport + 2, column=3).value = femaleTotalIns
        sheetReport.cell(row=maxRowReport + 1, column=4).value = maleTotalReg
        sheetReport.cell(row=maxRowReport + 2, column=4).value = femaleTotalReg
        # sheetReport.cell(row=maxRowReport + 1, column=7).value = maleWeekIns
        # sheetReport.cell(row=maxRowReport + 2, column=7).value = femaleWeekIns
        sheetReport.cell(row=maxRowReport + 1, column=7).value = maleWeekReg
        sheetReport.cell(row=maxRowReport + 2, column=7).value = femaleWeekReg

        # move some data
        counterFrow = 4
        counterT1 = 65
        counterT2 = 115
        i = 0
        while counterFrow + i < 51:
            sheetReport.cell(row=counterT1 + i, column=1).value = sheetReport.cell(row=counterFrow + i, column=1).value
            sheetReport.cell(row=counterT2 + i, column=1).value = sheetReport.cell(row=counterFrow + i, column=1).value
            sheetReport.cell(row=counterT1 + i, column=2).value = sheetReport.cell(row=counterFrow + i, column=2).value
            sheetReport.cell(row=counterT2 + i, column=2).value = sheetReport.cell(row=counterFrow + i, column=2).value
            sheetReport.cell(row=counterT1 + i, column=3).value = sheetReport.cell(row=counterFrow + i, column=4).value
            sheetReport.cell(row=counterT2 + i, column=3).value = sheetReport.cell(row=counterFrow + i, column=7).value
            i += 1
        sheetReport.cell(row=114, column=3).value = sheetReport.cell(row=2, column=7).value
        print(">>100!>>>")
        bookReport.save("TemplateReport.xlsx")

    # bookReport.save("Отчет по регистрациям в Мой Доктор на " + currStamp + ".xlsx")
def checkUniqueClient():
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    maxRowBase = sheetBase.max_row
    counterE = 3
    print("Processing Base... ", end="")
    display = 0
    while counterE <= maxRowBase:
        if counterE / maxRowBase * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 3
        if sheetBase.cell(row=counterE, column=20).value == None:
            sheetBase.cell(row=counterE, column=20).value = "uniq"
            counterDuplicate = counterE - 1
            while counterDuplicate > 1:
                if sheetBase.cell(row=counterE, column=13).value == sheetBase.cell(row=counterDuplicate,
                                                                                   column=13).value:
                    sheetBase.cell(row=counterE, column=20).value = "rept"
                    break
                if sheetBase.cell(row=counterE, column=3).value == sheetBase.cell(row=counterDuplicate,
                                                                                  column=3).value and sheetBase.cell(
                        row=counterDuplicate, column=20).value == "uniq":
                    if sheetBase.cell(row=counterE, column=13).value != sheetBase.cell(row=counterDuplicate,
                                                                                       column=13).value:
                        sheetBase.cell(row=counterE, column=20).value = "uniq"
                        sheetBase.cell(row=counterDuplicate, column=20).value = "rept"
                    else:
                        sheetBase.cell(row=counterE, column=20).value = "rept"
                counterDuplicate -= 1
        counterE += 1
    print(">>100!>>>")
    bookBase.save(pathBase)
def checkActive(smsReport: str):
    bookSMS = openpyxl.load_workbook(smsReport)
    sheetSMS = bookSMS.active
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    maxRowBase = sheetBase.max_row
    maxRowSMS = sheetSMS.max_row
    counterE = 3
    print("Processing Base... ", end="")
    display = 0
    while counterE <= maxRowBase:
        if counterE / maxRowBase * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 3
        if sheetBase.cell(row=counterE, column=21).value != "aktiv":
            sheetBase.cell(row=counterE, column=21).value = "neakt"
            counterSMS = 2
            while counterSMS <= maxRowSMS:
                if sheetBase.cell(row=counterE, column=3).value == sheetSMS.cell(row=counterSMS, column=3).value:
                    sheetBase.cell(row=counterE, column=21).value = "aktiv"
                    break
                counterSMS += 1
        counterE += 1
    print("")
    bookBase.save(pathBase)
def extractTop100():
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    while (input("00 to EXIT: ") != "00"):
        counterClients = 1
        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        currWeek = week_num
        maxRow = sheetBase.max_row
        randomNum = random.randrange(1, 99)
        callerName = input("Call-Center Name: ")
        callerPref1 = input("Desired office 1: ")
        # callerPref2 = input("Desired office 2: ")
        callerPref2 = callerPref1
        f = open("список-" + callerName + "-" + currStamp + "-" + str(randomNum) + ".txt", 'w', encoding='utf-8')
        f.write("список- " + callerName + " " + str(currTime))
        f.write('\n')
        newList = True
        match = False
        listSize = 240
        if callerPref1 != "":
            listSize = 360
        counterA = maxRow - 1
        while counterA > 1 and counterClients < listSize:
            if callerPref1 in sheetBase.cell(row=counterA, column=2).value or callerPref2 in sheetBase.cell(
                    row=counterA, column=2).value:
                match = True
            else:
                match = False
            if sheetBase.cell(row=counterA, column=25).value == None and sheetBase.cell(row=counterA,
                                                                                        column=20).value == "uniq" and sheetBase.cell(
                    row=counterA, column=21).value == "neakt" and match:
                f.write(
                    str(counterClients) + " " + sheetBase.cell(row=counterA, column=3).value + " / " + sheetBase.cell(
                        row=counterA, column=2).value[5:] + " +" + sheetBase.cell(
                        row=counterA, column=13).value)
                f.write('\n')
                sheetBase.cell(row=counterA, column=25).value = callerName
                sheetBase.cell(row=counterA, column=24).value = currWeek
                newList = True
                counterClients += 1
            counterA -= 1
            if counterClients % 30 == 0 and newList is True:
                f.write('\n')
                f.write('---------------')
                f.write('\n')
                f.write("список- " + callerName + " " + str(currTime))
                f.write('\n')
                newList = False
        f.close()
        print("List of contacts created")
    bookBase.save(pathBase)
def extractRepeats():
    bookBase = openpyxl.load_workbook(pathBase)
    sheetBase = bookBase.active
    while (input("00 to EXIT: ") != "00"):
        counterClients = 1
        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        currWeek = week_num
        maxRow = sheetBase.max_row
        randomNum = random.randrange(1, 99)
        callerName = input("PREV Call-Center Name: ")
        callerPref1 = ""
        # callerPref1 = input("Desired office: ")
        callFinal = callerName
        # callFinal = input("NEW Call-Center Name: ")
        f = open("список-" + callFinal + "-" + currStamp + "- R -" + str(randomNum) + ".txt", 'w', encoding='utf-8')
        f.write("список- " + callFinal + " R  из " + callerName + " " + str(currTime))
        f.write('\n')
        newList = True
        match = False
        listSize = 360
        counterA = maxRow - 1
        match = False
        while counterA > 1 and counterClients < listSize:
            if callerPref1 in sheetBase.cell(row=counterA, column=2).value:
                match = True
            else:
                match = False
            if sheetBase.cell(row=counterA, column=25).value == callerName and sheetBase.cell(row=counterA,
                                                                                              column=20).value == "uniq" and sheetBase.cell(
                    row=counterA, column=21).value == "neakt" and match:
                f.write(
                    str(counterClients) + " " + sheetBase.cell(row=counterA, column=3).value + " / " + sheetBase.cell(
                        row=counterA, column=2).value[5:] + " +" + sheetBase.cell(
                        row=counterA, column=13).value)
                f.write('\n')
                sheetBase.cell(row=counterA, column=25).value = "R " + callFinal
                sheetBase.cell(row=counterA, column=26).value = callerName
                sheetBase.cell(row=counterA, column=24).value = currWeek
                newList = True
                counterClients += 1
            counterA -= 1
            if counterClients % 30 == 0 and newList is True:
                f.write('\n')
                f.write('---------------')
                f.write('\n')
                f.write("список- " + callFinal + " R  из " + callerName + " " + str(currTime))
                f.write('\n')
                newList = False
        f.close()
        print("List of contacts created")
    bookBase.save(pathBase)
def transformDocList(docList: str):
    bookList = openpyxl.load_workbook(docList)
    sheetList = bookList.active
    bookNew = openpyxl.load_workbook("TemplateDocs.xlsx")
    sheetNew = bookNew.active

    maxRowList = sheetList.max_row
    counterA = 6
    counterNew = 1
    while counterA <= maxRowList:
        # print("start")
        while sheetList.cell(row=counterA, column=7).value == None and counterA <= maxRowList:
            counterA += 1
        counterB = counterA - 1
        while counterB > 0:
            # print(str(counterA) + " / " + str(counterB) + str(sheetList.cell(row=counterA, column=3).value) + str(sheetNew.cell(row=counterB, column=2).value))
            if (sheetList.cell(row=counterA, column=1).value == sheetNew.cell(row=counterB, column=1).value and
                    sheetList.cell(row=counterA, column=3).value == sheetNew.cell(row=counterB, column=2).value and
                    sheetList.cell(row=counterA, column=7).value == sheetNew.cell(row=counterB, column=4).value):
                counterA += 1
                counterB = counterA - 1
                if counterA > maxRowList:
                    break
            else:
                counterB -= 1
        if counterA > maxRowList:
            break

        counterColumn = 1
        for i in [1, 3, 6, 7]:
            sheetNew.cell(row=counterNew, column=counterColumn).value = sheetList.cell(row=counterA, column=i).value
            counterColumn += 1
        insCase = str(sheetNew.cell(row=counterNew, column=4).value)
        sheetNew.cell(row=counterNew, column=5).value = "заявление"
        sheetNew.cell(row=counterNew, column=6).value = "ожидаем"
        if "смерть" in insCase:
            counterNew += 1
            for i in [1, 2, 3, 4, 6]:
                sheetNew.cell(row=counterNew, column=i).value = sheetNew.cell(row=counterNew - 1, column=i).value
            sheetNew.cell(row=counterNew, column=5).value = "свидетельство о смерти нот.зав.копия"
        elif "инвалидность" in insCase:
            counterNew += 1
            for i in [1, 2, 3, 4, 6]:
                sheetNew.cell(row=counterNew, column=i).value = sheetNew.cell(row=counterNew - 1, column=i).value
            sheetNew.cell(row=counterNew, column=5).value = "нот.зав.копия МСЭК"
        elif "критическое" in insCase:
            for i in [1, 2, 3, 4, 6]:
                sheetNew.cell(row=counterNew, column=i).value = sheetNew.cell(row=counterNew - 1, column=i).value
            sheetNew.cell(row=counterNew, column=5).value = "эпикриз оригинал"
        if sheetNew.cell(row=counterNew, column=4).value is None:
            counterNew -= 1
        counterNew += 1
        counterA += 1

    bookNew.save(
        f"БФоригиналы-{str(sheetNew.cell(row=1, column=1).value)[:5]}-{str(sheetNew.cell(row=counterNew - 1, column=1).value)[:5]}.xlsx")
    # print(sheetNew.max_row)
    print(f"DOCS Report - {docList} - Processed")
def prepeareSKReport():
    bookAll = openpyxl.load_workbook(pathAll)
    sheetAll = bookAll.active
    bookReport = openpyxl.load_workbook(pathRepSK)
    sheetReport = bookReport.active
    maxRowAll = sheetAll.max_row
    print("Preparing SK REPORT... ", end="")
    display = 0
    currMonth = str(sheetAll.cell(row=maxRowAll, column=1).value)[:8]
    counterMain = 2
    counterReport = 2
    while counterMain < maxRowAll:
        qThisM = 0
        havePolicy = sheetAll.cell(row=counterMain, column=7).value != "НЕТ ПОЛИСА"
        planned = sheetAll.cell(row=counterMain, column=11).value != "дежурный"
        newCard = sheetAll.cell(row=counterMain, column=9).value == "открытие файла"
        thisMonth = str(sheetAll.cell(row=counterMain, column=1).value)[:8] == currMonth
        if thisMonth and havePolicy and newCard:
            sheetReport.cell(row=counterReport, column=1).value = sheetAll.cell(row=counterMain, column=1).value
            sheetReport.cell(row=counterReport, column=2).value = sheetAll.cell(row=counterMain, column=5).value
            sheetReport.cell(row=counterReport, column=3).value = "открытие файла"
            sheetReport.cell(row=counterReport, column=4).value = sheetAll.cell(row=counterMain, column=7).value
            sheetReport.cell(row=counterReport, column=5).value = sheetAll.cell(row=counterMain, column=8).value
            sheetReport.cell(row=counterReport, column=6).value = sheetAll.cell(row=counterMain, column=9).value
            sheetReport.cell(row=counterReport, column=7).value = sheetAll.cell(row=counterMain, column=10).value
            sheetReport.cell(row=counterReport, column=8).value = sheetAll.cell(row=counterMain, column=11).value
            counterReport += 1

        if thisMonth and havePolicy and planned:
            counter = maxRowAll
            while counter > 1:
                if sheetAll.cell(row=counterMain, column=5).value == sheetAll.cell(row=counter, column=5).value \
                        and str(sheetAll.cell(row=counter, column=1).value)[:8] == currMonth \
                        and sheetAll.cell(row=counter, column=11).value == "плановый":
                    qThisM += 1
                counter -= 1
            checkDuplicate = sheetReport.max_row
            while checkDuplicate > 1:
                if sheetReport.cell(row=checkDuplicate, column=2).value == sheetAll.cell(row=counterMain,
                                                                                         column=5).value \
                        and sheetReport.cell(row=checkDuplicate, column=3).value == "специалист":
                    break
                checkDuplicate -= 1

            if checkDuplicate == 1:
                sheetReport.cell(row=counterReport, column=1).value = sheetAll.cell(row=counterMain, column=1).value
                sheetReport.cell(row=counterReport, column=2).value = sheetAll.cell(row=counterMain, column=5).value
                sheetReport.cell(row=counterReport, column=3).value = "специалист"
                sheetReport.cell(row=counterReport, column=4).value = sheetAll.cell(row=counterMain, column=7).value
                sheetReport.cell(row=counterReport, column=5).value = sheetAll.cell(row=counterMain, column=8).value
                sheetReport.cell(row=counterReport, column=6).value = sheetAll.cell(row=counterMain, column=9).value
                sheetReport.cell(row=counterReport, column=7).value = sheetAll.cell(row=counterMain, column=10).value
                sheetReport.cell(row=counterReport, column=8).value = sheetAll.cell(row=counterMain, column=11).value
                sheetReport.cell(row=counterReport, column=9).value = qThisM
                counterReport += 1

        if thisMonth and havePolicy and not planned:
            counter = maxRowAll
            while counter > 1:
                if sheetAll.cell(row=counterMain, column=5).value == sheetAll.cell(row=counter, column=5).value \
                        and str(sheetAll.cell(row=counter, column=1).value)[:8] == currMonth \
                        and sheetAll.cell(row=counter, column=11).value == "дежурный":
                    qThisM += 1
                counter -= 1
            checkDuplicate = sheetReport.max_row
            while checkDuplicate > 1:
                if sheetReport.cell(row=checkDuplicate, column=2).value == sheetAll.cell(row=counterMain,
                                                                                         column=5).value \
                        and sheetReport.cell(row=checkDuplicate, column=3).value == "дежурный":
                    break
                checkDuplicate -= 1

            if checkDuplicate == 1:
                sheetReport.cell(row=counterReport, column=1).value = sheetAll.cell(row=counterMain, column=1).value
                sheetReport.cell(row=counterReport, column=2).value = sheetAll.cell(row=counterMain, column=5).value
                sheetReport.cell(row=counterReport, column=3).value = "дежурный"
                sheetReport.cell(row=counterReport, column=4).value = sheetAll.cell(row=counterMain, column=7).value
                sheetReport.cell(row=counterReport, column=5).value = sheetAll.cell(row=counterMain, column=8).value
                sheetReport.cell(row=counterReport, column=6).value = sheetAll.cell(row=counterMain, column=9).value
                sheetReport.cell(row=counterReport, column=7).value = sheetAll.cell(row=counterMain, column=10).value
                sheetReport.cell(row=counterReport, column=8).value = sheetAll.cell(row=counterMain, column=11).value
                sheetReport.cell(row=counterReport, column=9).value = qThisM
                counterReport += 1

        counterMain += 1

        if counterMain / maxRowAll * 100 > display:
            print(">" + str(display) + "%", end="")
            display += 3
    sheetReport.delete_cols(6)
    bookReport.save("SKReport.xlsx")
def attendance(path_xlsx):
    all_date = []
    path_json = 'resourses/file_name.json'
    empty_dict = {
        'Name': [],
        'DateTime': [],
        'Door': []}
    all_df = []
    matching = []
    plan_list = {"data": []}
    # Create the DataFrame
    # end_dataframe = pd.DataFrame(empty_dict)
    xlsx_frame = pd.read_excel(path_xlsx, header=None, skiprows=[0, 1]).rename(
        columns={1: 'Name', 3: 'DateTime', 5: 'Door'})
    xlsx_frame.head()
    all_date = xlsx_frame["DateTime"].unique()
    for i, data in enumerate(all_date):
        all_date[i] = data[0:10]
    date_set = sorted(set(all_date))

    for date in date_set:
        query_in = f"""SELECT Name, DateTime, Door FROM xlsx_frame WHERE Door LIKE '%doorIN%' AND DateTime LIKE '%{date}%'
                                                            GROUP BY Name
                                                            ORDER BY DateTime """.format(date)
        new_data_frame = sqldf(query_in, locals())
        query_out = f"""SELECT Name, DateTime, Door FROM xlsx_frame WHERE Door LIKE '%doorOUT%' AND DateTime LIKE '%{date}%'
                                                            GROUP BY Name
                                                            ORDER BY DateTime """.format(date)
        data_out = sqldf(query_out, locals())
        for data in data_out.Name.values:
            if data not in new_data_frame.Name.values:
                not_exit = {
                    'date': date,
                    'name': data
                }
                matching.append(not_exit)
        all_df.append(new_data_frame)
    end_dataframe = pd.concat(all_df)
    end_dataframe.to_excel("Attendance.xlsx", index=False)

    data_json = end_dataframe.to_json(orient='records', force_ascii=False)

    with open(path_json, 'w', encoding='utf8') as f:
        f.write(data_json)
        # json.dump(out, f, indent=2, ensure_ascii=False)
    f = open(path_json, encoding="utf8'")
    new_data = json.load(f)

    for data in new_data:

        date = data["DateTime"][0:10]
        worker = data["Name"]
        first_income = data["DateTime"][11:19]
        query = f"""SELECT MAX(DateTime) FROM xlsx_frame WHERE Door LIKE '%doorOUT%' AND DateTime LIKE '%{date}%' 
                                                                    AND Name LIKE '%{worker}%'""".format(date, worker)
        last_exit_query = sqldf(query, locals())

        if last_exit_query.values:
            last_exit = last_exit_query.values[0][0][11:19]

        else:
            last_exit = ''

        new_value = {
            "Дата": date,
            "Сотрудник": worker,
            "Первый приход": first_income,
            "Последний уход": last_exit,
        }
        plan_list["data"].append(new_value)
    for match in matching:
        date = match["date"]
        worker = match['name']
        query = f"""SELECT MAX(DateTime) FROM xlsx_frame WHERE Door LIKE '%doorOUT%' AND DateTime LIKE '%{date}%' 
                                                                            AND Name LIKE '%{worker}%'""".format(date,
                                                                                                                 worker)
        last_exit_query = sqldf(query, locals())
        new_value = {
            "Дата": date,
            "Сотрудник": worker,
            "Первый приход": '',
            "Последний уход": last_exit_query.values[0][0][11:19],
        }
        plan_list["data"].append(new_value)
    name_xlsx = 'attendance.xlsx'
    with open('resourses/attendance.json', 'w', encoding='utf8') as f:
        json.dump(plan_list, f, indent=2, ensure_ascii=False)
    pd.read_json("resourses/attendance.json", orient='split').sort_values('Дата').to_excel(name_xlsx, index=False)
    print("Создана таблица с посещением")
def extractFree():
    bookBase = openpyxl.load_workbook(pathFree)
    sheetBase = bookBase.active
    while (input("00 to EXIT: ") != "00"):
        counterClients = 1
        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        currWeek = week_num
        maxRow = sheetBase.max_row
        randomNum = random.randrange(1, 99)
        callerName = input("Call-Center Name: ")
        f = open("AGсписок-" + callerName + "-" + currStamp + "-" + str(randomNum) + ".txt", 'w', encoding='utf-8')
        f.write("AGсписок- " + callerName + " " + str(currTime))
        f.write('\n')
        newList = True
        match = False
        listSize = 120
        counterA = maxRow - 1
        while sheetBase.cell(row=counterA, column=7).value != None:
            counterA -= 1
        while counterA > 1 and counterClients < listSize:
            match = True
            if sheetBase.cell(row=counterA, column=7).value == None:
                f.write(
                    str(counterClients) + " " + sheetBase.cell(row=counterA,
                                                               column=1).value + " / рег от " + sheetBase.cell(
                        row=counterA, column=4).value + " +" + sheetBase.cell(
                        row=counterA, column=2).value)
                f.write('\n')
                sheetBase.cell(row=counterA, column=7).value = callerName
                sheetBase.cell(row=counterA, column=8).value = currWeek
                newList = True
                counterClients += 1
            counterA -= 5
            if counterClients % 30 == 0 and newList is True:
                f.write('\n')
                f.write('---------------')
                f.write('\n')
                f.write("AGсписок- " + callerName + " " + str(currTime))
                f.write('\n')
                newList = False
        f.close()
        print("List of contacts created")
    bookBase.save(pathFree)
def dailyReceptions():
    for file in glob.glob("rece*.xlsx"):
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Delete columns
        ws.delete_cols(17)
        ws.delete_cols(13)
        ws.delete_cols(12)
        ws.delete_cols(11)
        ws.delete_cols(7)
        ws.delete_cols(7)
        print(f"Rows after deleting columns in {file}: {ws.max_row}")  # Checkpoint 1

        # Filter rows
        counter = ws.max_row
        while counter > 1:
            if ws.cell(row=counter, column=7).value == None:
                ws.cell(row=counter, column=7).value = "---"
            if "Исакова Айпери" in ws.cell(row=counter, column=5).value or "Рысбекова Зарина Рысбековна" in ws.cell(
                    row=counter, column=5).value or ws.cell(row=counter, column=8).value == "дежурный":
                ws.delete_rows(counter)
            elif ws.cell(row=counter, column=9).value == None and ws.cell(row=counter, column=10).value == None:
                ws.cell(row=counter, column=9).value = "НЕУСПЕШНЫЙ"
                ws.cell(row=counter, column=10).value = "НЕУСПЕШНЫЙ"
            counter -= 1
        print(f"Rows after filtering in {file}: {ws.max_row}")  # Checkpoint 2

        # Check third column values before sorting
        for i in range(1, ws.max_row + 1):
            print(ws.cell(row=i, column=3).value)  # Checkpoint 3

        value_counts = {}  # Dictionary to hold counts of each unique value in column 3
        for i in range(1, ws.max_row + 1):
            value = ws.cell(row=i, column=3).value
            if value in value_counts:
                value_counts[value] += 1
            else:
                value_counts[value] = 1

        # Print out the counts
        print(f"\nOccurrences in {file}:")
        for value, count in value_counts.items():
            print(f"'{value}': {count} times")
        print("-" * 40)  # Print a separator for clarity

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        data_rows = rows[1:]

        sorted_data_rows = sorted(data_rows, key=lambda x: x[2])
        wb.remove(ws)
        ws_sorted = wb.create_sheet("SortedSheet", 0)

        for col_num, value in enumerate(header, 1):
            ws_sorted.cell(row=1, column=col_num, value=value)

        for row_num, row in enumerate(sorted_data_rows, 2):
            for col_num, value in enumerate(row, 1):
                ws_sorted.cell(row=row_num, column=col_num, value=value)
        ws_sorted.delete_cols(8)
        # ws_sorted.delete_cols(7)
        wb.save("proc-" + ws_sorted.cell(row=2, column=1).value + " " + file)
        os.remove(file)
def writeToLog(file):
    # Создаем упорядоченный словарь для хранения статистики по датам и врачам
    doctor_stats = OrderedDict()

    # Открываем файл
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # Проходимся по строкам файла, начиная со второй строки
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = row[0]
        doctor = row[2]
        reception_topic = row[7]

        # Проверяем, есть ли запись для этой даты в словаре
        if date not in doctor_stats:
            doctor_stats[date] = {}

        # Проверяем, есть ли запись для этого врача в словаре
        if doctor not in doctor_stats[date]:
            doctor_stats[date][doctor] = {"Успешные приемы": 0, "Неуспешные приемы": 0}

        # Подсчитываем успешные и неуспешные приемы
        if "НЕУСПЕШНЫЙ" in reception_topic:
            doctor_stats[date][doctor]["Неуспешные приемы"] += 1
        else:
            doctor_stats[date][doctor]["Успешные приемы"] += 1

    # Сортировка словаря по дате
    sorted_doctor_stats = OrderedDict(sorted(doctor_stats.items(), key=lambda x: x[0]))

    # Проверим наличие существующего лог-файла и откроем его, если он существует
    if os.path.isfile("LogReceptions.xlsx"):
        log_wb = openpyxl.load_workbook("LogReceptions.xlsx")
        log_ws = log_wb.active
    else:
        # Если лог-файл не существует, создадим его
        log_wb = openpyxl.Workbook()
        log_ws = log_wb.active
        log_ws.append(["Дата", "ФИО врача", "Успешные приемы", "Неуспешные приемы"])

    # Добавление данных в существующий лог-файл
    for date, doctors in sorted_doctor_stats.items():
        for doctor, stats in doctors.items():
            log_ws.append([date, doctor, stats["Успешные приемы"], stats["Неуспешные приемы"]])

    # Добавление разделительной строки
    log_ws.append(["-" * 10, "-" * 10, "-" * 10, "-" * 10, "-" * 10])
    # Сохранение обновленного лог-файла
    log_wb.save("LogReceptions.xlsx")
def UpdateDates(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        current_date = datetime.datetime.now().date()
        formatted_date = current_date.strftime('%Y-%m-%d')  # Преобразуем дату в строку формата "yyyy-mm-dd"

        for row in sheet.iter_rows(min_row=4, min_col=5):
            cell = row[0]
            cell.value = formatted_date  # Записываем отформатированную дату как текст

        workbook.save(filename)
        print("Даты успешно обновлены.")

    except Exception as e:
        print(f"Произошла ошибка: {e}")
def process_excel_files_cost():
    # Открываем файлы
    log_receptions_file = "LogReceptions.xlsx"
    md_sotrudniki_file = "МД Сотрудники.xlsx"
    expenses_file = "PaymentsDaily.xlsx"

    # Открываем файлы Excel
    log_receptions_workbook = openpyxl.load_workbook(log_receptions_file)
    md_sotrudniki_workbook = openpyxl.load_workbook(md_sotrudniki_file)
    expenses_workbook = openpyxl.load_workbook(expenses_file)

    # Выбираем активные листы в каждой книге
    log_receptions_sheet = log_receptions_workbook.active
    md_sotrudniki_sheet = md_sotrudniki_workbook.active
    expenses_sheet = expenses_workbook.active

    # Проходимся по файлу "LogReceptions" и обновляем стоимости, если они не были ранее посчитаны
    counterLog = log_receptions_sheet.max_row
    maxRowSotr = md_sotrudniki_sheet.max_row
    while counterLog > 2:
        name = log_receptions_sheet.cell(row=counterLog, column=2).value  # Имя находится во 2 столбце
        cost = log_receptions_sheet.cell(row=counterLog, column=5).value  # Стоимость находится в 5 столбце
        if cost == None:
            payment = "N/A"
            finalP = "N/A"
            agent = "N/A"
            counterSotr = 2
            while md_sotrudniki_sheet.cell(row=counterSotr, column=5).value != None:
                nameS = md_sotrudniki_sheet.cell(row=counterSotr, column=5).value  # Имя находится в 5 столбце
                costS = md_sotrudniki_sheet.cell(row=counterSotr, column=18).value  # Стоимость находится в 18 столбце
                if name in nameS:
                    if md_sotrudniki_sheet.cell(row=counterSotr, column=21).value == "день":
                        finalP = costS * log_receptions_sheet.cell(row=counterLog,
                                                                   column=3).value + 150 * log_receptions_sheet.cell(
                            row=counterLog, column=4).value
                    else:
                        finalP = md_sotrudniki_sheet.cell(row=counterSotr, column=21).value
                    payment = md_sotrudniki_sheet.cell(row=counterSotr, column=20).value
                    agent = md_sotrudniki_sheet.cell(row=counterSotr, column=17).value
                    break
                counterSotr += 1
            log_receptions_sheet.cell(row=counterLog, column=5).value = finalP
            log_receptions_sheet.cell(row=counterLog, column=6).value = payment
            log_receptions_sheet.cell(row=counterLog, column=7).value = agent
            if finalP != "месяц":
                today = datetime.date.today()
                formatted_date = today.strftime('%d.%m.%Y')
                expenses_sheet.append(["З/П", \
                                       log_receptions_sheet.cell(row=counterLog, column=1).value, \
                                       log_receptions_sheet.cell(row=counterLog, column=2).value, \
                                       log_receptions_sheet.cell(row=counterLog, column=3).value, \
                                       log_receptions_sheet.cell(row=counterLog, column=4).value, \
                                       log_receptions_sheet.cell(row=counterLog, column=5).value, \
                                       log_receptions_sheet.cell(row=counterLog, column=6).value, \
                                       "квитанция эл.кошелек", \
                                       formatted_date])
        counterLog -= 1

    expenses_sheet.append(["-" * 10, "-" * 10, "-" * 10, "-" * 10, "-" * 10, "-" * 10, "-" * 10, "-" * 10])
    log_receptions_workbook.save("LogReceptions.xlsx")
    expenses_workbook.save(expenses_file)
    # Закрываем файлы
    log_receptions_workbook.close()
    md_sotrudniki_workbook.close()
    expenses_workbook.close()
def summarize_log():
    # Загрузка данных из файла Excel
    data = pd.read_excel('LogReceptions.xlsx', engine='openpyxl')

    # Группировка по имени (второй столбец) и суммирование значений по третьему, четвертому и пятому столбцам
    grouped_data = data.groupby(data.columns[1]).agg({
        data.columns[2]: 'sum',
        data.columns[3]: 'sum',
        data.columns[4]: 'sum',
        data.columns[6]: 'sum'
    }).reset_index()

    # Переименование столбцов для читаемости
    grouped_data.columns = ['Имя', 'Успешные приемы', 'Неуспешные приемы', 'Уплаченная сумма', 'Агент']

    # Сохранение результата в новый Excel-файл
    grouped_data.to_excel('ProcessedLogReceptions.xlsx', index=False, engine='openpyxl')

# ----------------M---A---I---N----------------

pathBase = "base.xlsx"
pathFree = "free.xlsx"
pathLog = "resourses/LOG.txt"
pathAll = "baseAllCalls.xlsx"
pathEmer = "resourses/baseEmer.xlsx"
pathPlan = "resourses/basePlan.xlsx"
pathOrg = "resourses/listOrg.txt"
pathSov = "resourses/listSov.txt"
pathBaseNew = "baseNew.xlsx"
pathPRCD = "PRCD-AllEntriesBF.xlsx"
pathRepSK = "resourses/TemplateSKReport.xlsx"
pathRepINS = "resourses/TemplateINS.xlsx"
currStamp = str(datetime.datetime.now())[:10]
currTime = (str(datetime.datetime.now())[5:16])

with open(pathLog, "a+") as f:
    f.write(f"{str(datetime.datetime.now())}\n")
    f.close()

while True:
    x = input(
        "------------------- \n11 for ОБРАБОТКА СПИСКОВ ЗАСТРАХОВАННЫХ \n22 for ТАБЕЛЬ ПОСЕЩЕНИЙ \n33 for ИМПОРТ СМС ОТЧЕТА / СПИСКА ПРИЕМОВ / ФОРМИРОВАНИЕ ОТЧЕТА"
        "\n44 for СПИСОК ДОКОВ СК \n55 for ТЕЛЕГРАМ ЧАТЫ \n66(6) for СПИСОК ОБЗВОНОВ \n77 for АГЕТСКИЕ \n88 for ЕЖЕДНЕВНЫЕ ПРИЕМЫ \n99 for ОБРАБОТКА ЛОГА \n00 for ВЫХОД \nIN: ")

    if x == "11":
        dailyImport()

    if x == "22":
        for file in glob.glob("report*.xlsx"):
            attendance(file)

    elif x == "33":
        for file in glob.glob("rece*.xlsx"):
            dateRange = updateCallsBase(file)
            source_path = file
            destination_path = "arhvRecep/recep" + dateRange + ".xlsx"
            new_location = shutil.move(source_path, destination_path)
            print("File {0} Processed and Moved to \n  > > > > >  {1}".format(source_path, new_location))
        prepeareSKReport()
        smsReportImport()
        checkActive(pathPRCD)

    elif x == "44":
        for file in glob.glob("доки*.xlsx"):
            transformDocList(file)

    elif x == "55":
        json_receptions = "plan.json"
        json_call = "emer.json"
        json_record = "record.json"
        TelegramEntries(json_record)
        TelegramScheldule(json_receptions)
        TelegramCall(json_call)
        TelegramCallCC()
        TelegramPlanCC()
        UpdateAgentList()

    elif x == "66":
        backupBase()
        extractTop100()

    elif x == "666":
        backupBase()
        extractRepeats()

    elif x == "77":
        extractFree()

    elif x == "88":
        backupLog()
        for file in glob.glob("proc-*.xlsx"):
            os.remove(file)
        dailyReceptions()
        for file in glob.glob("proc-*.xlsx"):
            writeToLog(file)
        UpdateDates("aaImport/abonement.xlsx")
        process_excel_files_cost()

    elif x == "99":
        summarize_log()


    elif x == "00":
        break

    else:
        print("--- NOT VALID ---")

print("--- COMPLETED ---")